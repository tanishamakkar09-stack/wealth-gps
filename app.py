from flask import (
    Flask,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    redirect,
    url_for,
    make_response
)
from flask_mail import Mail, Message
from datetime import datetime
import sqlite3
from weasyprint import HTML
import tempfile
import os
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
from werkzeug.utils import secure_filename


app = Flask(__name__) 
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.secret_key = app.config["SECRET_KEY"] 
import os

app.config["SECRET_KEY"] = os.getenv("SECRET_KEY")
app.secret_key = app.config["SECRET_KEY"]

app.config["ADMIN_USERNAME"] = os.getenv("ADMIN_USERNAME")
app.config["ADMIN_PASSWORD"] = os.getenv("ADMIN_PASSWORD")

app.config["MAIL_SERVER"] = os.getenv("MAIL_SERVER")
app.config["MAIL_PORT"] = int(os.getenv("MAIL_PORT"))
app.config["MAIL_USE_TLS"] = os.getenv("MAIL_USE_TLS") == "True"
app.config["MAIL_USERNAME"] = os.getenv("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.getenv("MAIL_PASSWORD")
mail = Mail(app)
EXCEL_FILE = "Meeting_Bookings.xlsx"
UPLOAD_FOLDER = "uploads"
CAS_FOLDER = os.path.join(UPLOAD_FOLDER, "CAS")
CAMS_FOLDER = os.path.join(UPLOAD_FOLDER, "CAMS")
REPORT_FOLDER = "reports"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CAS_FOLDER, exist_ok=True)
os.makedirs(CAMS_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

os.makedirs(REPORT_FOLDER, exist_ok=True)

os.makedirs(CAS_FOLDER, exist_ok=True)
os.makedirs(CAMS_FOLDER, exist_ok=True)


def save_booking_to_excel(data):

    if os.path.exists(EXCEL_FILE):

        wb = load_workbook(EXCEL_FILE)

        ws = wb.active

    else:

        wb = Workbook()

        ws = wb.active

        ws.title = "Bookings"

        ws.append([
            "Booking Date",
            "Client Name",
            "Mobile",
            "Email",
            "Occupation",
            "Marital Status",
            "Meeting Date",
            "Meeting Time",
            "Meeting Mode",
            "Risk Profile",
            "Annual Income",
            "Monthly SIP",
            "Financial Stability",
            "Retirement Readiness",
            "Goal Achievement",
            "Family Members",
            "CAS Uploaded",
            "CAMS Uploaded",
            "CAS File",
            "CAMS File",
            "Remarks",
            "Status",
            "PDF Report"
        ])

    ws.append([

        data.get("booking_date"),

        data.get("name"),

        data.get("mobile"),

        data.get("email"),

        data.get("occupation"),

        data.get("marital_status"),

        data.get("meeting_date"),

        data.get("meeting_time"),

        data.get("meeting_mode"),

        data.get("risk_profile"),

        data.get("annual_income"),

        data.get("monthly_sip"),

        data.get("financial_stability"),

        data.get("readiness"),

        data.get("goal_achievement"),

        data.get("family_count"),

        data.get("cas_uploaded"),

        data.get("cams_uploaded"),

        data.get("cas_filename"),

        data.get("cams_filename"),

        data.get("remarks"),
        "Pending",
        data.get("pdf_report")
    ])

    wb.save(EXCEL_FILE)

# ---------------- DATABASE ---------------- #

def get_connection():
    return sqlite3.connect("database.db")

def generate_charts(report):

    chart_folder = os.path.join("static", "charts")
    os.makedirs(chart_folder, exist_ok=True)

    # =====================================================
    # PORTFOLIO ALLOCATION PIE CHART
    # =====================================================

    plt.figure(figsize=(6,6))

    values = [
        report["equity_pct"],
        report["debt_pct"],
        report["gold_pct"],
        report["cash_pct"]
    ]

    labels = [
        "Equity",
        "Debt",
        "Gold",
        "Cash"
    ]

    plt.pie(
        values,
        labels=labels,
        autopct="%1.1f%%",
        startangle=90
    )

    plt.title("Portfolio Allocation")

    allocation_path = os.path.join(
        chart_folder,
        "allocation.png"
    )

    plt.savefig(
        allocation_path,
        bbox_inches="tight"
    )

    plt.close()


    # =====================================================
    # GOAL PROGRESS BAR CHART
    # =====================================================

    plt.figure(figsize=(7,4))

    goal_names = []

    goal_values = []

    goal_names.append(report["goal1_name"])
    goal_values.append(report["goal1_achievement"])

    if report["goal2_name"]:
        goal_names.append(report["goal2_name"])
        goal_values.append(report["goal2_achievement"])

    if report["goal3_name"]:
        goal_names.append(report["goal3_name"])
        goal_values.append(report["goal3_achievement"])

    plt.bar(goal_names, goal_values)

    plt.ylim(0,100)

    plt.ylabel("Achievement %")

    plt.title("Goal Progress")

    goal_path = os.path.join(
        chart_folder,
        "goal_progress.png"
    )

    plt.savefig(
        goal_path,
        bbox_inches="tight"
    )

    plt.close()


    return {

        "allocation_chart": allocation_path.replace("\\","/"),

        "goal_chart": goal_path.replace("\\","/")

    }
# ---------------- HOME ---------------- #

@app.route('/')
def home():
    return render_template('index.html')


# ---------------- FINANCIAL INPUTS ---------------- #

@app.route('/financial-inputs', methods=['POST'])
def financial_inputs():

    return render_template(
        'financial_inputs.html',
        name=request.form.get('name'),
        age=request.form.get('age'),
        mobile=request.form.get('mobile'),
        email=request.form.get('email'),
        portfolio=request.form.get('portfolio'),
        occupation=request.form.get("occupation"),
        marital_status=request.form.get("marital_status"),
    )

# ---------------- RISK PROFILE ---------------- #

@app.route('/risk-profile', methods=['POST'])
def risk_profile():

    return render_template(
        'risk_profile.html',

        # Client Details
        name=request.form.get('name'),
        mobile=request.form.get('mobile'),
        email=request.form.get('email'),
        occupation=request.form.get("occupation"),
        marital_status=request.form.get("marital_status"),

        # Financial Details
        age=request.form.get('age'),
        annual_income=request.form.get('annual_income'),
        monthly_sip=request.form.get('monthly_sip'),
        retirement_age=request.form.get('retirement_age'),
        dependents=request.form.get('dependents'),

        # Assets
        direct_equity=request.form.get("direct_equity"),
        mutual_funds=request.form.get("mutual_funds"),
        fd=request.form.get("fd"),
        ppf=request.form.get("ppf"),
        epf=request.form.get("epf"),
        nps=request.form.get("nps"),
        bonds=request.form.get("bonds"),
        gold=request.form.get("gold"),
        real_estate=request.form.get("real_estate"),
        insurance=request.form.get("insurance"),
        alternate=request.form.get("alternate"),
        cash=request.form.get("cash"),
        )

# ---------------- GOALS ---------------- #

@app.route('/goals', methods=['POST'])
def goals():

    return render_template(
        'goals.html',

        # Client Details
        name=request.form.get('name'),
        mobile=request.form.get('mobile'),
        email=request.form.get('email'),
        occupation=request.form.get("occupation"),
        marital_status=request.form.get("marital_status"),

        # Financial Details
        age=request.form.get('age'),
        annual_income=request.form.get('annual_income'),
        monthly_sip=request.form.get('monthly_sip'),
        retirement_age=request.form.get('retirement_age'),
        dependents=request.form.get('dependents'),

        # Assets
        direct_equity=request.form.get("direct_equity"),
        mutual_funds=request.form.get("mutual_funds"),
        fd=request.form.get("fd"),
        ppf=request.form.get("ppf"),
        epf=request.form.get("epf"),
        nps=request.form.get("nps"),
        bonds=request.form.get("bonds"),
        gold=request.form.get("gold"),
        real_estate=request.form.get("real_estate"),
        insurance=request.form.get("insurance"),
        alternate=request.form.get("alternate"),
        cash=request.form.get("cash"),

        # Risk Assessment
        return_expectation=request.form.get('return_expectation'),
        time_horizon=request.form.get('time_horizon'),
        investment_knowledge=request.form.get('investment_knowledge'),
        financial_situation=request.form.get('financial_situation'),
        portfolio_choice=request.form.get('portfolio_choice'),
    )

@app.route('/family-information', methods=['POST'])
def family_information():

    return render_template(

        'family_information.html',

        **request.form

    )

@app.route('/calculate', methods=['POST'])
def calculate():

    # =====================================================
    # CLIENT DETAILS
    # =====================================================

    name = request.form.get("name")
    mobile = request.form.get("mobile")
    email = request.form.get("email")

    occupation = request.form.get("occupation")
    marital_status = request.form.get("marital_status")


    # =====================================================
    # PERSONAL DETAILS
    # =====================================================

    age = int(request.form.get("age") or 30)

    annual_income = float(
        request.form.get("annual_income") or 0
    )

    monthly_sip = float(
        request.form.get("monthly_sip") or 0
    )

    retirement_age = int(
        request.form.get("retirement_age") or 60
    )

    dependents = int(
        request.form.get("dependents") or 0
    )

    years = retirement_age - age

    if years <= 0:
        years = 1


    # =====================================================
    # FINANCIAL PROFILE
    # =====================================================

    direct_equity = float(
        request.form.get("direct_equity") or 0
    )

    mutual_funds = float(
        request.form.get("mutual_funds") or 0
    )

    fd = float(
        request.form.get("fd") or 0
    )

    ppf = float(
        request.form.get("ppf") or 0
    )

    epf = float(
        request.form.get("epf") or 0
    )

    nps = float(
        request.form.get("nps") or 0
    )

    bonds = float(
        request.form.get("bonds") or 0
    )

    gold = float(
        request.form.get("gold") or 0
    )

    real_estate = float(
        request.form.get("real_estate") or 0
    )

    insurance = float(
        request.form.get("insurance") or 0
    )

    alternate = float(
        request.form.get("alternate") or 0
    )

    cash = float(
        request.form.get("cash") or 0
    )


        # =====================================================
    # OPTIONAL DOCUMENTS
    # =====================================================

    cas_statement = request.files.get("cas_statement")

    cams_statement = request.files.get("cams_statement")

    

    # =====================================================
    # SAVE UPLOADED DOCUMENTS
    # =====================================================

    # =====================================================
    # OPTIONAL DOCUMENTS
    # =====================================================



    # =====================================================
    # SAVE UPLOADED DOCUMENTS
    # =====================================================

    cas_filename = ""

    cams_filename = ""

    if cas_statement and cas_statement.filename:

        original_name = secure_filename(cas_statement.filename)

        cas_filename = f"{name}_CAS_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_name}"
        cas_statement.save(
            os.path.join(CAS_FOLDER, cas_filename)
        )

    if cams_statement and cams_statement.filename:

        original_name = secure_filename(cams_statement.filename)

        cams_filename = f"{name}_CAMS_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{original_name}"

        cams_statement.save(
            os.path.join(CAMS_FOLDER, cams_filename)
        )

    # =====================================================
    # ASSET CALCULATIONS
    # =====================================================fsessio
    # =====================================================
    # ASSET CALCULATIONS
    # =====================================================

    equity = direct_equity + mutual_funds

    debt = (

        fd +

        ppf +

        epf +

        nps +

        bonds

    )

    total_assets = (

        equity +

        debt +

        gold +

        real_estate +

        insurance +

        alternate +

        cash

    )

    if total_assets == 0:

        total_assets = 1


    equity_pct = round(

        (equity / total_assets) * 100,

        1

    )

    debt_pct = round(

        (debt / total_assets) * 100,

        1

    )

    gold_pct = round(

        (gold / total_assets) * 100,

        1

    )

    cash_pct = round(

        (cash / total_assets) * 100,

        1

    )
        # =====================================================
    # RISK PROFILE
    # =====================================================

    risk_score = (

        int(request.form.get("return_expectation") or 1)

        +

        int(request.form.get("time_horizon") or 1)

        +

        int(request.form.get("investment_knowledge") or 1)

        +

        int(request.form.get("financial_situation") or 1)

        +

        int(request.form.get("portfolio_choice") or 1)

    )

    if risk_score <= 10:

        risk_profile = "Conservative"

    elif risk_score <= 18:

        risk_profile = "Moderate"

    else:

        risk_profile = "Aggressive"


    # =====================================================
    # RETIREMENT CALCULATIONS
    # =====================================================

    annual_return = 0.15

    inflation_rate = 0.07

    monthly_return = annual_return / 12


    corpus = round(

        monthly_sip *

        (

            (

                (1 + monthly_return)

                **

                (years * 12)

                - 1

            )

            /

            monthly_return

        )

    )


    delayed_years = max(

        years - 5,

        1

    )


    delayed_corpus = round(

        monthly_sip *

        (

            (

                (1 + monthly_return)

                **

                (delayed_years * 12)

                - 1

            )

            /

            monthly_return

        )

    )


    cost_of_delay = corpus - delayed_corpus


    # =====================================================
    # SAVINGS RATE
    # =====================================================

    if annual_income > 0:

        savings_rate = (

            (monthly_sip * 12)

            /

            annual_income

        ) * 100

    else:

        savings_rate = 0


    # =====================================================
    # RETIREMENT READINESS
    # =====================================================

    expected_retirement_corpus = max(

        annual_income * 20,

        10000000

    )

    readiness = round(

        min(

            (corpus / expected_retirement_corpus) * 100,

            100

        )

    )


    # =====================================================
    # WEALTH SCORE
    # =====================================================

    wealth_score = round(

        (

            readiness * 0.50

        )

        +

        (

            min(savings_rate, 50) * 0.30

        )

        +

        (

            (100 - cash_pct) * 0.20

        )

    )

    wealth_score = min(

        wealth_score,

        100

    )


    # =====================================================
    # FINANCIAL STABILITY
    # =====================================================

    financial_stability = round(

        (

            min(savings_rate, 50) * 40

            / 50

        )

        +

        (

            readiness * 0.30

        )

        +

        (

            (100 - cash_pct) * 0.20

        )

        +

        (

            (10 - min(dependents, 10)) * 2

        )

    )

    financial_stability = min(

        financial_stability,

        100
    )
        # =====================================================
    # GOALS
    # =====================================================

    current_year = datetime.now().year

    # ---------------- GOAL 1 ---------------- #

    goal1 = request.form.get("goal1")

    goal1_other = request.form.get("goal1_other")

    if goal1 == "Other (Please Specify)":

        goal1_name = goal1_other.strip()

    else:

        goal1_name = goal1


    goal1_corpus = float(

        request.form.get("goal1_corpus") or 0

    )

    goal1_year = int(

        request.form.get("goal1_year") or current_year

    )


    # ---------------- GOAL 2 ---------------- #

    goal2 = request.form.get("goal2")

    goal2_other = request.form.get("goal2_other")

    if goal2 == "Other (Please Specify)":

        goal2_name = goal2_other.strip()

    else:

        goal2_name = goal2


    goal2_corpus = float(

        request.form.get("goal2_corpus") or 0

    )

    goal2_year = int(

        request.form.get("goal2_year") or current_year

    )


    # ---------------- GOAL 3 ---------------- #

    goal3 = request.form.get("goal3")

    goal3_other = request.form.get("goal3_other")

    if goal3 == "Other (Please Specify)":

        goal3_name = goal3_other.strip()

    else:

        goal3_name = goal3


    goal3_corpus = float(

        request.form.get("goal3_corpus") or 0

    )

    goal3_year = int(

        request.form.get("goal3_year") or current_year

    )


    # =====================================================
    # GOAL CALCULATOR
    # =====================================================

    def calculate_goal(goal_amount, goal_year):

        years_left = max(

            goal_year - current_year,

            1

        )

        future_value = round(

            goal_amount *

            ((1 + inflation_rate) ** years_left)

        )

        projected_value = round(

            monthly_sip *

            (

                (

                    (1 + monthly_return)

                    **

                    (years_left * 12)

                    - 1

                )

                /

                monthly_return

            )

        )

        achievement = round(

            min(

                (projected_value / future_value) * 100,

                100

            )

        )

        return (

            future_value,

            projected_value,

            achievement,

            years_left

        )


    goal1_future,\
    goal1_projection,\
    goal1_achievement,\
    goal1_years = calculate_goal(

        goal1_corpus,

        goal1_year

    )


    if goal2_corpus > 0:

        goal2_future,\
        goal2_projection,\
        goal2_achievement,\
        goal2_years = calculate_goal(

            goal2_corpus,

            goal2_year

        )

    else:

        goal2_future = 0

        goal2_projection = 0

        goal2_achievement = None

        goal2_years = 0


    if goal3_corpus > 0:

        goal3_future,\
        goal3_projection,\
        goal3_achievement,\
        goal3_years = calculate_goal(

            goal3_corpus,

            goal3_year

        )

    else:

        goal3_future = 0

        goal3_projection = 0

        goal3_achievement = None

        goal3_years = 0


    # =====================================================
    # OVERALL GOAL ACHIEVEMENT
    # =====================================================

    goal_scores = [

        goal1_achievement

    ]

    if goal2_achievement is not None:

        goal_scores.append(

            goal2_achievement

        )

    if goal3_achievement is not None:

        goal_scores.append(

            goal3_achievement

        )

    goal_achievement = round(

        sum(goal_scores)

        /

        len(goal_scores)

    )


    target_corpus = goal1_corpus

    future_target_corpus = goal1_future

    target_year = goal1_year

    years_to_goal = goal1_years
        # =====================================================
    # ADVISOR INSIGHTS
    # =====================================================

    advisor_comment = ""

    if wealth_score >= 85:

        advisor_comment += (
            "Your overall financial position is strong with excellent long-term wealth creation potential. "
            "Continue reviewing your portfolio periodically and increase your SIP annually."
        )

    elif wealth_score >= 70:

        advisor_comment += (
            "You have built a healthy financial foundation. "
            "Increasing investments gradually and maintaining diversification can further strengthen your portfolio."
        )

    elif wealth_score >= 50:

        advisor_comment += (
            "Your financial plan is progressing reasonably well, but additional investments and regular reviews are recommended."
        )

    else:

        advisor_comment += (
            "Your current financial preparedness requires improvement. "
            "A disciplined investment strategy and higher monthly savings are recommended."
        )


    # =====================================================
    # ASSET ALLOCATION INSIGHT
    # =====================================================

    if equity_pct > 80:

        advisor_comment += (

            "\n\n• Equity exposure is relatively high. "
            "Consider increasing allocation towards fixed-income investments to reduce volatility."

        )

    elif debt_pct > 70:

        advisor_comment += (

            "\n\n• Debt allocation is comparatively high. "
            "A gradual increase in equity allocation may improve long-term returns."

        )

    elif cash_pct > 25:

        advisor_comment += (

            "\n\n• Idle cash holdings appear significant. "
            "Investing surplus funds systematically could improve wealth creation."

        )


    # =====================================================
    # DEPENDENTS INSIGHT
    # =====================================================

    if dependents >= 4:

        advisor_comment += (

            "\n\n• You have multiple financial dependents. "
            "Ensure adequate emergency reserves, life insurance and succession planning."

        )

    elif dependents >= 2:

        advisor_comment += (

            "\n\n• Future family responsibilities should be incorporated while reviewing investments."

        )


    # =====================================================
        # ================= FAMILY INFORMATION =================

    family_count = int(request.form.get("family_count") or 0)

    family_members = []

    for i in range(1, family_count + 1):

        member_name = request.form.get(f"member_name_{i}", "").strip()

        member_relation = request.form.get(f"member_relation_{i}", "").strip()

        member_age = request.form.get(f"member_age_{i}", "").strip()

        member_occupation = request.form.get(f"member_occupation_{i}", "").strip()

        if not (member_name or member_relation or member_age or member_occupation):

            continue

        family_members.append({

            "name": member_name,

            "relation": member_relation,

            "age": member_age,

            "occupation": member_occupation

        })
    # =====================================================
    # START SAVING REPORT
    # =====================================================

    session["report"] = {

        # CLIENT
        "name": name,
        "mobile": mobile,
        "email": email,
        "occupation": occupation,
        "marital_status": marital_status,

        # PERSONAL
        "age": age,
        "annual_income": annual_income,
        "monthly_sip": monthly_sip,
        "retirement_age": retirement_age,
        "dependents": dependents,

        # FAMILY
        "family_count": family_count,
        "family_members": family_members,

        # FINANCIAL PROFILE
        "direct_equity": direct_equity,
        "mutual_funds": mutual_funds,
        "fd": fd,
        "ppf": ppf,
        "epf": epf,
        "nps": nps,
        "bonds": bonds,
        "gold": gold,
        "real_estate": real_estate,
        "insurance": insurance,
        "alternate": alternate,
        "cash": cash,

        # DOCUMENTS

"cas_uploaded": bool(cas_statement),

"cams_uploaded": bool(cams_statement),

"cas_filename": cas_filename if cas_filename else "Not Uploaded",

"cams_filename": cams_filename if cams_filename else "Not Uploaded",

        # ASSET SUMMARY
        "equity": equity,
        "debt": debt,
        "equity_pct": equity_pct,
        "debt_pct": debt_pct,
        "gold_pct": gold_pct,
        "cash_pct": cash_pct,

        # SCORES
        "wealth_score": wealth_score,
        "financial_stability": financial_stability,
        "readiness": readiness,
        "risk_score": risk_score,
        "risk_profile": risk_profile,

        # RETIREMENT
        "corpus": f"{corpus:,.0f}",
        "corpus_numeric": corpus,
        "years": years,
        "delayed_corpus": f"{delayed_corpus:,.0f}",
        "cost_of_delay": f"{cost_of_delay:,.0f}",

        # GOALS
        "goal1_name": goal1_name,
        "goal1_corpus": f"{goal1_corpus:,.0f}",
        "goal1_year": goal1_year,
        "goal1_achievement": goal1_achievement,

        "goal2_name": goal2_name,
        "goal2_corpus": f"{goal2_corpus:,.0f}",
        "goal2_year": goal2_year,
        "goal2_achievement": goal2_achievement,

        "goal3_name": goal3_name,
        "goal3_corpus": f"{goal3_corpus:,.0f}",
        "goal3_year": goal3_year,
        "goal3_achievement": goal3_achievement,

        "goal_achievement": goal_achievement,
        "target_corpus": f"{round(target_corpus):,}",
        "future_target_corpus": f"{round(future_target_corpus):,}",
        "target_year": target_year,
        "years_to_goal": years_to_goal,

        # REPORT
        "report_date": datetime.now().strftime("%d %B %Y"),
        "report_id": "ARP-" + datetime.now().strftime("%Y%m%d-%H%M%S"),
        "advisor_comment": advisor_comment
    }

    # ================= GENERATE DASHBOARD CHARTS =================

    charts = generate_charts(session["report"])

    session["report"].update(charts)

    session.modified = True

    return render_template(
        "dashboard.html",
        **session["report"]
    )
# ---------------- REVIEW BOOKING PAGE ---------------- #

@app.route('/review-booking')
def review_booking():

    today = datetime.today().strftime("%Y-%m-%d")
  
    return render_template(
        'review_booking.html',
        today=today
    )
@app.route("/download-report", methods=["POST"])
def download_report():

    report = session.get("report")

    if not report:
        return "No report data found. Please complete the assessment first."

    charts = generate_charts(report)
    report.update(charts)

    html = render_template(
        "report.html",
        **report
    )

    filename = f'Anand_Rathi_Preferred_Financial_Report_{report["name"]}.pdf'

    pdf_path = os.path.join(
        REPORT_FOLDER,
        filename
    )

    try:
        print("Generating PDF...")

        HTML(
            string=html,
            base_url=os.path.abspath(".")
        ).write_pdf(pdf_path)

        print("PDF Generated Successfully")

    except Exception as e:
        print("PDF Error:", e)
        raise

    return send_file(
        pdf_path,
        as_attachment=True,
        download_name=filename
    )

# ---------------- BOOK REVIEW ---------------- #
@app.route('/book-review', methods=['POST'])
def book_review():

    name = request.form.get('name')
    mobile = request.form.get('mobile')
    email = request.form.get('email')

    relationship_manager = request.form.get('relationship_manager')

    meeting_date = request.form.get('meeting_date')
    meeting_time = request.form.get('meeting_time')
    notes = request.form.get('notes')

    # ================= SAVE MEETING DETAILS =================

    session["report"]["relationship_manager"] = relationship_manager
    session["report"]["meeting_date"] = meeting_date
    session["report"]["meeting_time"] = meeting_time
    session["report"]["notes"] = notes

    session.modified = True

    from datetime import date

    if meeting_date < str(date.today()):
        return "Past dates are not allowed."

    # -------- Send Email (Optional) -------- #

    print("Email temporarily disabled")
            # =====================================================
        # =====================================================
    # SAVE BOOKING TO EXCEL
    # =====================================================

    report = session.get("report", {})

    try:

        save_booking_to_excel({

            "booking_date": datetime.now().strftime("%d-%m-%Y %H:%M"),

            "name": name,

            "mobile": mobile,

            "email": email,

            "occupation": report.get("occupation"),

            "marital_status": report.get("marital_status"),

            "meeting_date": meeting_date,

            "meeting_time": meeting_time,

            "meeting_mode": notes,

            "risk_profile": report.get("risk_profile"),

            "annual_income": report.get("annual_income"),

            "monthly_sip": report.get("monthly_sip"),

            "financial_stability": report.get("financial_stability"),

            "readiness": report.get("readiness"),

            "goal_achievement": report.get("goal_achievement"),

            "family_count": report.get("family_count"),

            "cas_uploaded": "Yes" if report.get("cas_uploaded") else "No",

            "cams_uploaded": "Yes" if report.get("cams_uploaded") else "No",

            "cas_filename": report.get("cas_filename", "Not Uploaded"),

            "cams_filename": report.get("cams_filename", "Not Uploaded"),

            "remarks": notes,

            "pdf_report": f"Anand_Rathi_Preferred_Financial_Report_{name}.pdf"

        })

    except Exception as e:

        print("Excel Save Error:", e)
    # -------- Booking Success Page -------- #

    return render_template(

    "booking_success.html",

    name=name,
    mobile=mobile,
    email=email,

    relationship_manager=relationship_manager,

    meeting_date=meeting_date,
    meeting_time=meeting_time,

    notes=notes

)
   
# ---------------- SEND CONFIRMATION EMAIL ---------------- #

def send_confirmation_email(client_name, client_email, meeting_date, meeting_time, rm_name):

    msg = Message(

        subject="Portfolio Review Meeting Confirmation",

        recipients=[client_email]

    )

    msg.body = f"""
Dear {client_name},

Thank you for choosing Anand Rathi Preferred.

Your portfolio review meeting has been successfully scheduled.

Meeting Details:

Relationship Manager : {rm_name}

Meeting Date : {meeting_date}

Meeting Time : {meeting_time}

Our Relationship Manager will contact you shortly.

Thank you for choosing Anand Rathi Preferred.

Regards,

Anand Rathi Preferred.
"""

    try:
        mail.send(msg)
        print("Email sent successfully")

    except Exception as e:
        print("Email failed:", e)
        #-------------------CONTACT-------------------#
@app.route('/contact')
def contact():
    return render_template('contact.html')

# ---------------- ABOUT ---------------- #

@app.route('/about')
def about():
    return render_template('about.html')

# ---------------- ADMIN PANEL ---------------- #

@app.route('/admin-leads')
def admin_leads():

    conn = sqlite3.connect('database.db')

    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM leads"
    )

    leads = cursor.fetchall()

    conn.close()

    return render_template(
        'admin_leads.html',
        leads=leads
    )
# =====================================================
# RELATIONSHIP MANAGER PORTAL
# =====================================================

@app.route("/admin")
def admin_dashboard():
    if not session.get("admin_logged_in"):

     return redirect("/admin-login")

    if os.path.exists(EXCEL_FILE):

        df = pd.read_excel(EXCEL_FILE)

        clients = df.to_dict("records")

        total_clients = len(df)

        pending = (
            len(df[df["Status"] == "Pending"])
            if "Status" in df.columns else 0
        )
        confirmed = (
        len(df[df["Status"] == "Confirmed"])
        if "Status" in df.columns else 0
    )

        completed = (
        len(df[df["Status"] == "Completed"])
        if "Status" in df.columns else 0
    )

        reports = total_clients

        documents = 0

        if "CAS Uploaded" in df.columns:
            documents += (df["CAS Uploaded"] == "Yes").sum()

        if "CAMS Uploaded" in df.columns:
            documents += (df["CAMS Uploaded"] == "Yes").sum()

    else:

        clients = []

        total_clients = 0

        pending = 0
        confirmed = 0

        completed = 0

        reports = 0

        documents = 0

    return render_template(

    "admin_dashboard.html",

    clients=clients,

    total_clients=total_clients,

    pending=pending,

    confirmed=confirmed,

    completed=completed,

    reports=reports,

    documents=documents

)
# =====================================================
# UPDATE STATUS
# =====================================================

@app.route("/update-status/<int:index>", methods=["POST"])
def update_status(index):

    if not os.path.exists(EXCEL_FILE):
        return "Excel file not found."

    df = pd.read_excel(EXCEL_FILE)
    if "Status" not in df.columns:
        df["Status"] = "Pending"
    if index >= len(df):
        return "Invalid client."

    df.loc[index, "Status"] = request.form.get("status")

    df.to_excel(EXCEL_FILE, index=False)

    return redirect("/admin")
# =====================================================
# CLIENT DETAILS
# =====================================================

@app.route("/client/<int:index>")
def client_details(index):

    if not os.path.exists(EXCEL_FILE):

        return "No client data found."

    df = pd.read_excel(EXCEL_FILE)

    clients = df.to_dict("records")

    if index >= len(clients):

        return "Client not found."

    client = clients[index]

    return render_template(

        "client_details.html",

        client=client

    )
# =====================================================
# DOWNLOAD EXCEL
# =====================================================

@app.route("/download-excel")
def download_excel():

    if not os.path.exists(EXCEL_FILE):

        return "Excel file not found."

    return send_file(

        EXCEL_FILE,

        as_attachment=True,

        download_name="Meeting_Bookings.xlsx"

    )
# =====================================================
# DOWNLOAD SAVED PDF
# =====================================================

@app.route("/download-pdf/<filename>")
def download_saved_pdf(filename):

    file_path = os.path.join(REPORT_FOLDER, filename)

    if not os.path.exists(file_path):
        return "PDF Report not found."

    return send_file(
        file_path,
        as_attachment=True
    )
# =====================================================
# ADMIN LOGIN
# =====================================================

@app.route("/admin-login", methods=["GET", "POST"])
def admin_login():

    if request.method == "POST":

        username = request.form.get("username")
        password = request.form.get("password")

        if (
            username == app.config["ADMIN_USERNAME"]
            and
            password == app.config["ADMIN_PASSWORD"]
        ):

            session["admin_logged_in"] = True

            return redirect("/admin")

        return render_template(
            "admin_login.html",
            error="Invalid Username or Password"
        )

    return render_template("admin_login.html")
# =====================================================
# ADMIN LOGOUT
# =====================================================

@app.route("/logout")
def logout():

    session.pop("admin_logged_in", None)

    return redirect("/admin-login")
@app.route("/view-document/<folder>/<filename>")
def view_document(folder, filename):

    if folder == "cas":
        folder_path = CAS_FOLDER

    elif folder == "cams":
        folder_path = CAMS_FOLDER

    else:
        return "Invalid folder."

    file_path = os.path.join(folder_path, filename)

    if not os.path.exists(file_path):
        return "Document not found."

    return send_from_directory(folder_path, filename)
@app.route("/view-pdf/<filename>")
def view_pdf(filename):

    return send_from_directory(
        REPORT_FOLDER,
        filename
    )
# ---------------- RUN ---------------- #

if __name__ == '__main__':
    app.run(debug=False)